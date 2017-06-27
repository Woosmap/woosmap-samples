from openpyxl import load_workbook
import json
import os
import time
import requests
from hashlib import sha1

INPUT_EXCEL_FILE = 'Food Markets.xlsx'
WORKSHEET_NAME = 'foodmarkets'
WOOSMAP_PRIVATE_API_KEY = '23713926-1af5-4321-ba54-032966f6e95d'
BATCH_SIZE = 5


class WoosmapAPIHelper:
    """A wrapper around the Woosmap Data API."""

    WOOSMAP_API_HOSTNAME = 'api.woosmap.com'

    def __init__(self):
        self.session = requests.Session()

    def delete(self):
        self.session.delete('https://{hostname}/stores/'.format(hostname=self.WOOSMAP_API_HOSTNAME),
                            params={'private_key': WOOSMAP_PRIVATE_API_KEY})

    def post(self, payload):
        return self.session.post('https://{hostname}/stores/'.format(hostname=self.WOOSMAP_API_HOSTNAME),
                                 params={'private_key': WOOSMAP_PRIVATE_API_KEY},
                                 json={'stores': payload})

    def end(self):
        self.session.close()


class ExcelFile:
    """A simple wrapper around the needed openpyxl functions for this script"""

    def __init__(self, filename, worksheet_name=''):
        self.filename = filename
        self.workbook = load_workbook(self.filename)
        self.worksheet_name = worksheet_name if worksheet_name else self.get_first_worksheet_name()
        self.worksheet = self.workbook.get_sheet_by_name(self.worksheet_name)

    def get_first_worksheet_name(self):
        return self.workbook.get_sheet_names()[0]

    def iter_rows(self):
        for row in self.worksheet.iter_rows():
            yield [cell.value for cell in row]


def get_name(asset):
    name = asset.get('Name', '')
    if name:
        return name
    else:
        raise ValueError('Unable to get the Name')


def generate_id(asset):
    asset_id = sha1(get_name(asset).encode('utf-8')).hexdigest()
    return asset_id


def get_contact(asset):
    return {
        'website': asset.get('Website', ''),
        'phone': asset.get('Contact Phone', ''),
        'email': asset.get('Contact Email', '')
    }


def get_geometry(asset):
    latitude = asset.get('Latitude', None)
    longitude = asset.get('Longitude', None)
    if latitude is not None and longitude is not None:
        return {
            'lat': float(latitude),
            'lng': float(longitude)
        }
    else:
        raise ValueError('Unable to get the location')


def get_address(asset):
    return {
        'lines': [asset.get('Address Line', '')],
        'city': asset.get('City', ''),
        'zipcode': asset.get('Zipcode', '')
    }


def convert_to_woosmap(asset):
    converted_asset = {}
    try:
        converted_asset.update({
            'storeId': generate_id(asset),
            'name': get_name(asset),
            'address': get_address(asset),
            'contact': get_contact(asset),
            'location': get_geometry(asset)
        })
    except ValueError as ve:
        print('ValueError Raised {0} for Asset {1}'.format(ve, json.dumps(asset, indent=2)))

    return converted_asset


def import_assets(assets_data, woosmap_api_helper):
    try:
        print('Batch import {count} Assets...'.format(count=len(assets_data)))
        response = woosmap_api_helper.post(assets_data)
        if response.status_code >= 400:
            response.raise_for_status()

    except requests.exceptions.HTTPError as http_exception:
        if http_exception.response.status_code >= 400:
            print('Woosmap API Import Error: {0}'.format(http_exception.response.text))
        else:
            print('Error requesting the API: {0}'.format(http_exception))
        return False
    except Exception as exception:
        print('Failed importing Assets! {0}'.format(exception))
        return False

    print('Successfully imported in {0} seconds'.format(response.elapsed.total_seconds()))
    return True


def batch(assets_data, n=1):
    l = len(assets_data)
    for ndx in range(0, l, n):
        yield assets_data[ndx:min(ndx + n, l)]


def main():
    start = time.time()
    print('Start parsing and importing your data...')
    excel_file = ExcelFile(INPUT_EXCEL_FILE, WORKSHEET_NAME)
    sheet_data = list(excel_file.iter_rows())
    header = sheet_data.pop(0)
    assets_as_dict = [dict(zip(header, item)) for item in sheet_data]

    woosmap_assets = []
    for asset in assets_as_dict:
        converted_asset = convert_to_woosmap(asset)
        if converted_asset:
            woosmap_assets.append(converted_asset)

    print('{0} Assets converted from source file'.format(len(woosmap_assets)))

    woosmap_api_helper = WoosmapAPIHelper()
    # /!\ deleting existing assets before posting new ones /!\
    woosmap_api_helper.delete()

    count_imported_assets = 0
    for chunk in batch(woosmap_assets):
        imported_success = import_assets(chunk, woosmap_api_helper)
        if imported_success:
            count_imported_assets += len(chunk)

    woosmap_api_helper.end()

    end = time.time()
    print('...Script ended in {0} seconds'.format(end - start))


if __name__ == '__main__':
    file_path = os.path.join(os.getcwd(), INPUT_EXCEL_FILE)
    if os.path.exists(file_path):
        main()
    else:
        print('File not found: {0} '.format(file_path))
